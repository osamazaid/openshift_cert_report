import subprocess
import json
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill
import base64
from cryptography import x509
from cryptography.hazmat.backends import default_backend
import concurrent.futures
import os
import warnings # Keep this import

# --- Utility Functions ---

def run_oc_command(command, namespace=None, ignore_not_found=False):
    """
    Executes an oc command and returns its JSON output.
    Raises an exception if the command fails, unless ignore_not_found is True and the error is 'resource type not found'.
    """
    cmd_list = ["oc", *command]
    if namespace:
        cmd_list.extend(["-n", namespace])
    cmd_list.extend(["-o", "json"])

    try:
        result = subprocess.run(cmd_list, capture_output=True, text=True, check=True)
        return json.loads(result.stdout)
    except subprocess.CalledProcessError as e:
        if "not found" in e.stderr.lower():
            if ignore_not_found:
                return None
            else:
                print(f"Error: Resource not found for command: {' '.join(cmd_list)} in {namespace if namespace else 'cluster-wide'}")
                print(f"Stderr: {e.stderr}")
                return None
        elif "forbidden" in e.stderr.lower() or "denied" in e.stderr.lower():
            print(f"  Access denied for command: {' '.join(cmd_list)} in {namespace if namespace else 'cluster-wide'}. Skipping.")
            return None
        else:
            print(f"Error executing command: {' '.join(cmd_list)}")
            print(f"Stderr: {e.stderr}")
            raise

def get_certificate_details(pem_data):
    """
    Parses a PEM-encoded certificate to extract expiration date and issuer.
    """
    try:
        if isinstance(pem_data, str):
            pem_data = pem_data.encode('utf-8')
            
        cert = x509.load_pem_x509_certificate(pem_data, default_backend())
        expiration_date = cert.not_valid_after_utc
        
        # --- Critical Change: Suppress warning directly here ---
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning) # Ignore UserWarnings within this block
            issuer = cert.issuer.rfc4514_string()
        # -----------------------------------------------------
        
        # Truncate issuer string if too long for openpyxl cell attribute (still good practice)
        if len(issuer) > 64:
            issuer = issuer[:57] + "..." # Truncate to 60 chars (57 + 3 for ellipsis)
            
        return expiration_date, issuer
    except Exception as e:
        # print(f"Error parsing certificate: {e}") # Uncomment for debugging certificate parse errors
        return None, None

# --- Main Logic and Parallel Functions (no changes needed here) ---

def process_namespace_certificates(ns_name, cert_manager_installed_global):
    """
    Collects certificate information for a single namespace.
    This function will be run in parallel for different namespaces.
    Returns a list of certificate dictionaries for this namespace.
    """
    namespace_certificates = []
    
    print(f"Processing namespace: {ns_name}")

    # 1. Secrets of type kubernetes.io/tls
    try:
        secrets_data = run_oc_command(["get", "secrets", "--field-selector=type=kubernetes.io/tls"], namespace=ns_name)
        if secrets_data:
            for secret in secrets_data.get("items", []):
                if secret.get("data") and "tls.crt" in secret["data"]:
                    expiration, issuer = get_certificate_details(base64.b64decode(secret["data"]["tls.crt"]))
                    if expiration:
                        namespace_certificates.append({
                            "Namespace": ns_name,
                            "Name": secret["metadata"]["name"],
                            "Type": "Secret (kubernetes.io/tls)",
                            "Expiration Date": expiration,
                            "Issuer": issuer,
                        })
    except Exception:
        pass 

    # 2. OpenShift Routes
    try:
        routes_data = run_oc_command(["get", "routes"], namespace=ns_name)
        if routes_data:
            for route in routes_data.get("items", []):
                if route.get("spec", {}).get("tls", {}).get("certificate"):
                    cert_data_base64 = route["spec"]["tls"]["certificate"]
                    try:
                        expiration, issuer = get_certificate_details(base64.b64decode(cert_data_base64))
                        if expiration:
                            namespace_certificates.append({
                                "Namespace": ns_name,
                                "Name": route["metadata"]["name"],
                                "Type": "Route (embedded)",
                                "Expiration Date": expiration,
                                "Issuer": issuer,
                            })
                    except Exception:
                        pass 
    except Exception:
        pass

    # 3. Kubernetes Ingresses
    try:
        ingresses_data = run_oc_command(["get", "ingresses"], namespace=ns_name)
        if ingresses_data:
            for ingress in ingresses_data.get("items", []):
                if ingress.get("spec", {}).get("tls"):
                    for tls_spec in ingress["spec"]["tls"]:
                        if tls_spec.get("secretName"):
                            secret_name = tls_spec["secretName"]
                            try:
                                secret_data = run_oc_command(["get", "secret", secret_name], namespace=ns_name)
                                if secret_data and secret_data.get("type") == "kubernetes.io/tls" and secret_data.get("data") and "tls.crt" in secret_data["data"]:
                                    expiration, issuer = get_certificate_details(base64.b64decode(secret_data["data"]["tls.crt"]))
                                    if expiration:
                                        namespace_certificates.append({
                                            "Namespace": ns_name,
                                            "Name": f"{ingress['metadata']['name']} (Ingress TLS: {secret_name})",
                                            "Type": "Secret (via Ingress)",
                                            "Expiration Date": expiration,
                                            "Issuer": issuer,
                                        })
                            except Exception:
                                pass
    except Exception:
        pass

    # 4. Cert-Manager Certificates CRs
    if cert_manager_installed_global:
        certs_cr_data = run_oc_command(["get", "certificates.cert-manager.io"], namespace=ns_name, ignore_not_found=True)
        
        if certs_cr_data:
            for cert_cr in certs_cr_data.get("items", []):
                secret_name = cert_cr.get("spec", {}).get("secretName")
                if secret_name:
                    try:
                        secret_data = run_oc_command(["get", "secret", secret_name], namespace=ns_name)
                        if secret_data and secret_data.get("type") == "kubernetes.io/tls" and secret_data.get("data") and "tls.crt" in secret_data["data"]:
                            expiration, issuer = get_certificate_details(base64.b64decode(secret_data["data"]["tls.crt"]))
                            if expiration:
                                namespace_certificates.append({
                                    "Namespace": ns_name,
                                    "Name": f"{cert_cr['metadata']['name']} (Cert-Manager)",
                                    "Type": "Cert-Manager Certificate",
                                    "Expiration Date": expiration,
                                    "Issuer": issuer,
                                })
                    except Exception:
                        pass
    return namespace_certificates


def list_openshift_certificates_oc_client():
    """
    Lists certificates from OpenShift secrets, routes, ingresses, and cert-manager CRs using oc client.
    Processes namespaces in parallel to speed up collection.
    """
    all_certificates = []
    
    # --- Step 1: Check if Cert-Manager CRD exists (globally, once) ---
    print("Checking for Cert-Manager operator installation...")
    cert_manager_installed_global = False
    cert_manager_crd_check = run_oc_command(["get", "crd", "certificates.cert-manager.io"], ignore_not_found=True)
    if cert_manager_crd_check:
        cert_manager_installed_global = True
        print("Cert-Manager operator CRD 'certificates.cert-manager.io' found. Will query Cert-Manager resources.")
    else:
        print("Cert-Manager operator CRD 'certificates.cert-manager.io' not found. Skipping Cert-Manager resource queries.")


    # Get all namespaces (this still happens serially, as it's typically quick)
    try:
        namespaces_data = run_oc_command(["get", "namespaces"])
        namespaces = [ns["metadata"]["name"] for ns in namespaces_data.get("items", [])]
        if not namespaces:
            print("No namespaces found or accessible.")
            return []
    except Exception as e:
        print(f"Failed to get namespaces: {e}")
        return []

    # Determine optimal number of worker threads for ThreadPoolExecutor
    max_workers = min(os.cpu_count() * 2, 20) if os.cpu_count() else 10 
    print(f"Using {max_workers} worker threads for parallel namespace processing.")

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_namespace = {
            executor.submit(process_namespace_certificates, ns_name, cert_manager_installed_global): ns_name
            for ns_name in namespaces
        }

        for future in concurrent.futures.as_completed(future_to_namespace):
            ns_name = future_to_namespace[future]
            try:
                namespace_certs = future.result()
                all_certificates.extend(namespace_certs)
            except Exception as exc:
                print(f"Namespace {ns_name} generated an exception during processing: {exc}")

    return all_certificates

# --- Excel Reporting Function (no changes needed here, as the warning is now suppressed earlier) ---

def create_excel_report(certificates_data, filename="openshift_certificates_report.xlsx"):
    """
    Creates an Excel report with two sheets: all certificates and expiring certificates.
    """
    workbook = openpyxl.Workbook()

    # Sheet 1: All Certificates
    sheet_all = workbook.active
    sheet_all.title = "All Certificates"
    headers_all = ["Namespace", "Name", "Type", "Expiration Date", "Issuer", "Status", "Remaining Days"]
    sheet_all.append(headers_all)

    # Apply bold font to headers
    for cell in sheet_all[1]:
        cell.font = Font(bold=True)

    today = datetime.now()
    
    for cert in certificates_data:
        if cert["Expiration Date"].tzinfo is not None and cert["Expiration Date"].tzinfo.utcoffset(cert["Expiration Date"]) is not None:
             cert_expiration_naive = cert["Expiration Date"].replace(tzinfo=None)
        else:
             cert_expiration_naive = cert["Expiration Date"]

        remaining_days = (cert_expiration_naive - today).days
        status = "Valid"
        if remaining_days < 0:
            status = "Expired"
        elif remaining_days <= 14:
            status = "Expiring Soon"

        row = [
            cert["Namespace"],
            cert["Name"],
            cert["Type"],
            cert_expiration_naive.strftime("%Y-%m-%d %H:%M:%S"),
            cert["Issuer"],
            status,
            remaining_days
        ]
        
        # Removed the 'with warnings.catch_warnings()' block here, as it's now in get_certificate_details
        sheet_all.append(row) 

        # Highlighting logic remains the same
        if status == "Expired":
            for cell in sheet_all[sheet_all.max_row]:
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif status == "Expiring Soon":
            for cell in sheet_all[sheet_all.max_row]:
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Auto-adjust column widths for Sheet 1
    for column in sheet_all.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet_all.column_dimensions[column_name].width = adjusted_width


    # Sheet 2: Certificates Expiring Within Two Weeks
    sheet_expiring = workbook.create_sheet("Expiring Soon")
    headers_expiring = ["Namespace", "Name", "Type", "Expiration Date", "Issuer", "Remaining Days"]
    sheet_expiring.append(headers_expiring)

    # Apply bold font to headers
    for cell in sheet_expiring[1]:
        cell.font = Font(bold=True)

    two_weeks_from_now = today + timedelta(weeks=2)

    for cert in certificates_data:
        if cert["Expiration Date"].tzinfo is not None and cert["Expiration Date"].tzinfo.utcoffset(cert["Expiration Date"]) is not None:
             cert_expiration_naive = cert["Expiration Date"].replace(tzinfo=None)
        else:
             cert_expiration_naive = cert["Expiration Date"]

        if today <= cert_expiration_naive <= two_weeks_from_now:
            remaining_days = (cert_expiration_naive - today).days
            row = [
                cert["Namespace"],
                cert["Name"],
                cert["Type"],
                cert_expiration_naive.strftime("%Y-%m-%d %H:%M:%S"),
                cert["Issuer"],
                remaining_days
            ]
            # Removed the 'with warnings.catch_warnings()' block here
            sheet_expiring.append(row) 

            for cell in sheet_expiring[sheet_expiring.max_row]:
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Auto-adjust column widths for Sheet 2
    for column in sheet_expiring.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet_expiring.column_dimensions[column_name].width = adjusted_width

    workbook.save(filename)
    print(f"\nReport generated: {filename}")

# --- Main Execution Block ---

if __name__ == "__main__":
    print("Collecting OpenShift certificate information using 'oc' client (parallel mode)...")
    certs = list_openshift_certificates_oc_client()
    if certs:
        create_excel_report(certs)
    else:
        print("No certificates found or accessible.")
